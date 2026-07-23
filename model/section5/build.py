"""Build Section 5 artefacts: figures, tables, CSV exports.

Outputs land in figures/ alongside section5.md / .pdf in the parent.
"""

from __future__ import annotations

import csv
import sys
from collections import defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
FIG = HERE / "figures"
FIG.mkdir(parents=True, exist_ok=True)

# Calibration constants (inherited from §3.4 / Section 5.2)
BETA = 0.99
DELTA = 0.96


# ------------------------------------------------------------------
# 1. Millennium data: annual UK macro 1086-2016
# ------------------------------------------------------------------

MILLENNIUM_COLS = {
    # column index in 'A1. Headline series' (1-indexed)
    "year": 1,
    "real_gdp": 2,
    "nom_gdp": 23,
    "cpi": 41,
    "cpi_inflation": 42,
    "bank_rate": 45,
    "gilt_10y": 47,
    "consols": 48,
    "corp_bond_yield": 51,
    "TME": 67,            # Total Managed Expenditure
    "receipts": 69,
    "net_borrowing": 71,  # Public Sector Net Lending(+)/Borrowing(-)
    "debt_par": 73,       # UK Public Sector Debt, par value, calendar year end
    "cg_debt_par": 76,    # Central Gov Gross Debt, par value, financial year end
    "cg_debt_mv": 77,     # Central Gov Gross Debt, MARKET VALUE, financial year end
}

DATA_ROW_START = 9  # data begins around row 9 (header rows are 1-8)


def load_millennium_a1() -> pd.DataFrame:
    # pandas is much faster than openpyxl row-by-row on this 27MB file
    cols_used = sorted(set(MILLENNIUM_COLS.values()))
    # Read with no header so we can pick rows ourselves; then keep only
    # the columns we need.
    raw = pd.read_excel(RAW / "millennium_data.xlsx", sheet_name="A1. Headline series",
                        header=None, usecols=[c - 1 for c in cols_used])
    # raw.columns are 0..N matching cols_used order
    col_map = {c: i for i, c in enumerate(cols_used)}
    df = pd.DataFrame()
    for name, src_col in MILLENNIUM_COLS.items():
        df[name] = raw.iloc[:, col_map[src_col]]
    # Data rows: where 'year' is a valid year 1700..2030
    df["year"] = pd.to_numeric(df["year"], errors="coerce")
    df = df[df["year"].between(1700, 2030)].copy()
    df["year"] = df["year"].astype(int)
    df = df.set_index("year").sort_index()
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# ------------------------------------------------------------------
# 2. BoE yield curve monthly: 10y gilt and 10y OIS
# ------------------------------------------------------------------

def load_yield_curve_monthly(folder: Path, sheet_candidates) -> pd.DataFrame:
    """Concatenate yield-curve files reading whichever candidate sheet exists.

    Different BoE yield-curve files use different sheet names for the spot
    curve — 2009-2015 OIS uses '2. spot curve', 2016+ uses '4. spot curve '
    (trailing space), the GLC files use '4. spot curve' without trailing space.
    Pass a list of candidates; first match wins per file.
    """
    if isinstance(sheet_candidates, str):
        sheet_candidates = [sheet_candidates]
    parts = []
    for path in sorted(folder.glob("*.xlsx")):
        # Discover which sheet exists in this file
        xl = pd.ExcelFile(path)
        sheet = None
        for cand in sheet_candidates:
            for sn in xl.sheet_names:
                if sn.strip() == cand.strip():
                    sheet = sn
                    break
            if sheet:
                break
        if sheet is None:
            print(f"  skip {path.name}: none of {sheet_candidates} found in {xl.sheet_names}")
            continue
        try:
            raw = pd.read_excel(path, sheet_name=sheet, header=None)
        except Exception as e:
            print(f"  skip {path.name}: {e}")
            continue
        if raw.empty:
            continue
        # Row 4 (0-indexed 3) is "years:" with col labels (col B onward)
        years_row = raw.iloc[3, :].values
        # Find maturity columns (numeric)
        cols = {}
        for i, v in enumerate(years_row):
            if isinstance(v, (int, float)) and v == v:  # not NaN
                cols[i] = float(v)
        # Data rows start at index 5 (row 6 in 1-indexed)
        data = raw.iloc[5:].copy()
        # Date column is col 0
        dates = pd.to_datetime(data.iloc[:, 0], errors="coerce")
        keep = dates.notna()
        data = data.loc[keep]
        dates = dates.loc[keep]
        mat_data = {cols[i]: pd.to_numeric(data.iloc[:, i], errors="coerce").values
                    for i in cols}
        df = pd.DataFrame(mat_data, index=pd.DatetimeIndex(dates.values))
        parts.append(df)
    if not parts:
        return pd.DataFrame()
    return pd.concat(parts).sort_index()


# ------------------------------------------------------------------
# 3. §3.4 ω time series (already calibrated) — read from inputs_filled_v2 output
# ------------------------------------------------------------------

def load_ftpl_omega() -> pd.Series:
    """Read the §3.4 ω series from outputs_filled_v2.xlsx."""
    path = HERE.parent / "calibration" / "outputs_filled_v2.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["calibration"]
    # Find year and omega columns
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    iy = hdr.index("year")
    iw = hdr.index("omega")
    rows = []
    for r in range(2, ws.max_row + 1):
        y = ws.cell(row=r, column=iy + 1).value
        w = ws.cell(row=r, column=iw + 1).value
        if isinstance(y, (int, float)) and isinstance(w, (int, float)):
            rows.append((int(y), float(w)))
    wb.close()
    s = pd.Series(dict(rows)).sort_index()
    s.name = "omega_ftpl"
    return s


# ------------------------------------------------------------------
# 4. Compute swap-spread ω (yield-based identification)
# ------------------------------------------------------------------

def compute_swap_spread_omega(target_maturity_y: float = 10.0) -> pd.Series:
    """Annual mean of (gilt_10y - ois_10y) monthly spread, converted to ω.

    Conversion: at the perpetuity steady state Q^L = β/(1-βδ),
      ω ≈ (yield_fund - yield_observed) / yield_fund · (something)
    More simply, since ω = (Q^L - Q^L,fund)/Q^L,fund and dQ/dy ≈ -Q·D,
      ω ≈ -D · (y_obs - y_fund)  =  D · swap_spread  (when gilt yield < OIS)
    Using D ≈ 1/(1-βδ) ≈ 19.8 years.
    """
    glc = load_yield_curve_monthly(RAW / "glcnom", ["4. spot curve"])
    ois = load_yield_curve_monthly(RAW / "ois", ["2. spot curve", "4. spot curve"])
    if glc.empty or ois.empty:
        return pd.Series(dtype=float, name="omega_swap")

    # Find target maturity columns
    def pick(df, target):
        cols = [c for c in df.columns if isinstance(c, (int, float))]
        nearest = min(cols, key=lambda c: abs(c - target))
        return df[nearest]

    g = pick(glc, target_maturity_y)
    o = pick(ois, target_maturity_y)
    # Align indices to common monthly observations
    spread_pct = g - o  # in %  (gilt - OIS); typically negative when gilts have convenience premium
    spread_dec = spread_pct / 100.0
    # Annual mean
    spread_ann = spread_dec.groupby(spread_dec.index.year).mean()
    # Duration approx
    D = 1.0 / (1.0 - BETA * DELTA)  # 19.8
    omega = -D * spread_ann  # negative spread -> positive omega
    omega.name = "omega_swap"
    omega.index.name = "year"
    return omega


# ------------------------------------------------------------------
# 5. §5.2 steady-state mapping table
# ------------------------------------------------------------------

def build_steady_state_table(mil: pd.DataFrame) -> pd.DataFrame:
    """Compute reference-period averages for the 2015-2019 window."""
    ref = mil.loc[2015:2019]
    nom_gdp = ref["nom_gdp"].mean()
    debt_par = ref["debt_par"].mean()
    cg_debt_mv = ref["cg_debt_mv"].mean()
    net_borrow = ref["net_borrowing"].mean()
    # Build the table
    rows = [
        ("v̄", "Real debt market value / GDP",
         "Central Govt debt MV (Millennium A1 col 77) / Nominal GDP (col 23/24)",
         cg_debt_mv / nom_gdp),
        ("s̄/ȳ", "Primary balance / GDP",
         "Public Sector Net Borrowing (col 71) / Nominal GDP — note: includes interest",
         net_borrow / nom_gdp),
        ("s̄/v̄", "Primary balance / debt MV",
         "(Public Sector Net Borrowing) / Central Govt debt MV",
         net_borrow / cg_debt_mv),
        ("w_S, w_L", "Short+medium and long shares of debt MV",
         "DMO Annual Review 2024-25 (Section 3.4 inheritance)", None),
        ("χ̄", "Long-bond share of net issuance",
         "DMO Annual Review (long share of conventional gilts)", None),
        ("b̄^L / ȳ", "Long-bucket gilt MV / GDP",
         "DMO (15+ bucket) × proportionality — Section 3.4 best-guess",
         0.18),
        ("b̄^S / ȳ", "Short+medium gilt MV / GDP", "DMO 0-15y bucket / GDP", 0.37),
        ("λ̄^L V̄^liab / ȳ", "Captive-demand pool / GDP",
         "PPF Purple Book s179 DB liabilities + BoE Insurance Aggregate technical provisions, "
         "averaged over 2015-2019",
         None),
        ("ω̄", "Long-end captive-demand wedge",
         "§3.4 sample average 2010-2024 (literature-β calibration): ω̄ = 0.121", 0.121),
        ("R̄/v̄", "Rent / debt MV",
         "Implied by FTPL steady state: 1 - β - s̄/v̄ = 0.04", 0.04),
    ]
    df = pd.DataFrame(rows, columns=["symbol", "name", "source", "value"])
    return df


# ------------------------------------------------------------------
# 6. §5.1 historical debt-to-GDP and real-return table
# ------------------------------------------------------------------

def build_history_table(mil: pd.DataFrame) -> pd.DataFrame:
    """1945 / 1960 / 1980 snapshots + 1945-1980 averages."""
    years = [1945, 1960, 1980]
    rows = []
    for y in years:
        rec = mil.loc[y]
        debt_to_gdp_par = rec["debt_par"] / rec["nom_gdp"] if pd.notna(rec["debt_par"]) else np.nan
        debt_to_gdp_mv = rec["cg_debt_mv"] / rec["nom_gdp"] if pd.notna(rec["cg_debt_mv"]) else np.nan
        rows.append({
            "year": y,
            "nominal_gdp_GBPmn": rec["nom_gdp"],
            "debt_par_GBPmn": rec["debt_par"],
            "debt_par_over_gdp": debt_to_gdp_par,
            "cg_debt_mv_GBPmn": rec["cg_debt_mv"],
            "cg_debt_mv_over_gdp": debt_to_gdp_mv,
            "gilt_10y_pct": rec["gilt_10y"],
            "cpi_inflation_pct": rec["cpi_inflation"],
            "real_return_pct": (rec["gilt_10y"] - rec["cpi_inflation"]) if pd.notna(rec["gilt_10y"]) and pd.notna(rec["cpi_inflation"]) else np.nan,
        })
    df = pd.DataFrame(rows).set_index("year")
    # Averages over 1945-1980
    sub = mil.loc[1945:1980]
    avg = {
        "nominal_gdp_GBPmn": np.nan,
        "debt_par_GBPmn": np.nan,
        "debt_par_over_gdp": (sub["debt_par"] / sub["nom_gdp"]).mean(),
        "cg_debt_mv_GBPmn": np.nan,
        "cg_debt_mv_over_gdp": (sub["cg_debt_mv"] / sub["nom_gdp"]).mean(),
        "gilt_10y_pct": sub["gilt_10y"].mean(),
        "cpi_inflation_pct": sub["cpi_inflation"].mean(),
        "real_return_pct": (sub["gilt_10y"] - sub["cpi_inflation"]).mean(),
    }
    df.loc["1945-1980 avg"] = avg
    return df


# ------------------------------------------------------------------
# 7. §5.4 figures
# ------------------------------------------------------------------

def fig1_omega_overlay(ftpl: pd.Series, swap: pd.Series, out: Path) -> None:
    """Two-panel plot: FTPL ω on top, swap-spread ω on bottom. The two
    series are on very different scales (FTPL spikes to 1.0+ in 2020,
    swap stays in ±0.1), so an overlay distorts. Two panels share the x-axis.
    """
    fig, (a1, a2) = plt.subplots(2, 1, figsize=(11, 6), sharex=True,
                                  gridspec_kw={"hspace": 0.18})

    # Top panel: FTPL ω
    a1.plot(ftpl.index, ftpl.values, lw=1.7, marker="s", ms=5,
            color="#d62728", label="FTPL-identity ω (§3.4)")
    a1.axhline(0, color="0.6", lw=0.6)
    a1.set_ylabel("FTPL ω")
    a1.set_title("UK long-end wedge ω — two identification approaches")
    a1.grid(alpha=0.3)
    a1.legend(loc="upper left")

    # Bottom panel: swap-spread ω
    if not swap.empty:
        a2.plot(swap.index, swap.values, lw=1.7, marker="o", ms=4,
                color="#1f77b4", label="Swap-spread ω (10y gilt − OIS, BoE)")
    a2.axhline(0, color="0.6", lw=0.6)
    a2.set_ylabel("Swap-spread ω")
    a2.set_xlabel("year")
    a2.grid(alpha=0.3)
    a2.legend(loc="lower left")

    # Regime annotations on both panels
    for x, lbl in [(2008, "QE begins"), (2022.7, "mini-budget")]:
        for ax in (a1, a2):
            ax.axvline(x, color="0.7", lw=0.6, ls="--")
        a1.text(x, a1.get_ylim()[1] * 0.95, lbl, fontsize=8,
                ha="center", va="top", color="0.4")
    a1.set_xlim(2008, 2025)

    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight")
    fig.savefig(out.with_suffix(".pdf"), bbox_inches="tight")
    plt.close(fig)


def fig2_conv_yield_comparison(swap: pd.Series, out: Path) -> None:
    """Convenience yield (in bps) = swap-spread reading scaled out."""
    fig, ax = plt.subplots(figsize=(11, 5))
    # Convert ω to yield-space basis points via 1/Q_fund - 1/Q
    # yield_wedge ≈ (1/Q_fund) * ω/(1+ω). At Q_fund≈19.8 -> 1/Q_fund=0.0505.
    Qf_inv = 1.0 / (BETA / (1.0 - BETA * DELTA))
    bp_wedge = 10000 * Qf_inv * swap / (1.0 + swap)
    ax.plot(bp_wedge.index, bp_wedge.values, lw=1.8, marker="o", ms=4,
            color="#1f77b4", label="Yield-wedge implied by 10y gilt-OIS spread")
    ax.axhline(0, color="0.6", lw=0.6)
    # Literature benchmarks (horizontal bands)
    ax.axhspan(30, 100, alpha=0.10, color="#2ca02c",
               label="Greenwood-Hanson-Vayanos / V-V steady-state range (30-100bps)")
    ax.set_xlim(2010, 2025)
    ax.set_xlabel("year")
    ax.set_ylabel("yield wedge (bps)")
    ax.set_title("UK long-gilt yield wedge in bps, vs. literature benchmarks")
    ax.legend(loc="upper right")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(out, dpi=150, bbox_inches="tight")
    fig.savefig(out.with_suffix(".pdf"), bbox_inches="tight")
    plt.close(fig)


# ------------------------------------------------------------------
# 8. Subperiod statistics
# ------------------------------------------------------------------

def subperiod_stats(ftpl: pd.Series, swap: pd.Series) -> pd.DataFrame:
    """Compute mean / SD of each ω series by subperiod, plus their correlation
    in the overlap.
    """
    periods = {
        "liberalized 1996-2007":  range(1996, 2008),
        "post-crisis 2008-2019":  range(2008, 2020),
        "post-2020":              range(2020, 2025),
    }
    rows = []
    for label, yrs in periods.items():
        yrs = list(yrs)
        f_sub = ftpl.reindex(yrs).dropna()
        s_sub = swap.reindex(yrs).dropna()
        common = sorted(set(f_sub.index) & set(s_sub.index))
        corr = float(f_sub.reindex(common).corr(s_sub.reindex(common))) if len(common) >= 3 else np.nan
        rows.append({
            "period": label,
            "n_years": len(yrs),
            "ftpl_mean": f_sub.mean() if not f_sub.empty else np.nan,
            "ftpl_sd":   f_sub.std()  if not f_sub.empty else np.nan,
            "swap_mean": s_sub.mean() if not s_sub.empty else np.nan,
            "swap_sd":   s_sub.std()  if not s_sub.empty else np.nan,
            "corr":      corr,
        })
    return pd.DataFrame(rows).set_index("period")


# ------------------------------------------------------------------
# main
# ------------------------------------------------------------------

def main():
    print("Loading data...")
    mil = load_millennium_a1()
    print(f"  Millennium A1: {len(mil)} rows, years {mil.index.min()}-{mil.index.max()}")
    ftpl = load_ftpl_omega()
    print(f"  FTPL ω: {ftpl.index.min()}-{ftpl.index.max()}, {len(ftpl)} years")
    swap = compute_swap_spread_omega(target_maturity_y=10.0)
    print(f"  swap-spread ω: {swap.index.min()}-{swap.index.max() if not swap.empty else 'n/a'}, {len(swap)} years")

    # Tables
    print("\nBuilding §5.1 history table...")
    hist = build_history_table(mil)
    hist.to_csv(FIG / "tbl_history.csv")
    print(hist.to_string())

    print("\nBuilding §5.2 steady-state mapping table...")
    ss = build_steady_state_table(mil)
    ss.to_csv(FIG / "tbl_steady_state.csv")
    print(ss.to_string())

    print("\nBuilding §5.4 subperiod stats...")
    stats = subperiod_stats(ftpl, swap)
    stats.to_csv(FIG / "tbl_subperiod_stats.csv")
    print(stats.to_string())

    # Save series themselves
    ftpl.to_csv(FIG / "omega_ftpl.csv")
    swap.to_csv(FIG / "omega_swap.csv")

    # Figures
    print("\nBuilding figures...")
    fig1_omega_overlay(ftpl, swap, FIG / "fig1_omega_overlay.png")
    fig2_conv_yield_comparison(swap, FIG / "fig2_conv_yield_bps.png")
    print(f"  wrote {FIG}/fig1_omega_overlay.png")
    print(f"  wrote {FIG}/fig2_conv_yield_bps.png")

    print("\nDone.")


if __name__ == "__main__":
    main()
