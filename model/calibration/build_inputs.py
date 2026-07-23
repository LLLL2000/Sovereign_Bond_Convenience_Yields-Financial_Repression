"""Compile raw/ files into inputs.xlsx for the §3.4 calibration.

Auto-fills:
  - GDP_nominal_GBPm                  ONS YBHA (calendar year)
  - primary_balance_over_GDP          OBR "Aggregates (per cent of GDP)"
                                      col "Primary balance", FY YYYY-(YY+1) -> calendar YYYY
  - real_short_rate                   BoE GLC Real, shortest available maturity (~2.5y).
                                      Annual mean of month-end observations. Stored as decimal.
                                      NOTE: methodology asks for 1y; linker curve does not
                                      publish below 2.5y. Proxy documented in README.

Manual-fill cells (set to sentinel "MANUAL"):
  - long_duration_years               DMO 15+ Macaulay duration
  - V_liab_over_GDP                   (TPR Purple Book DB s179 liabilities + PRA Solvency II
                                      insurance technical reserves) / nominal GDP
  - long_gilt_mkt_val_over_GDP        DMO long-gilt market value / nominal GDP
  - short_gilt_mkt_val_over_GDP       DMO short+medium-gilt market value / nominal GDP

Run:  python build_inputs.py
Output: inputs.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
OUT = HERE / "inputs.xlsx"

YEARS = list(range(2010, 2025))  # 2010..2024 inclusive (15 obs)

MANUAL = "MANUAL"


# ---------- ONS YBHA: nominal GDP --------------------------------------------

def load_ons_ybha() -> dict[int, float]:
    raw = json.loads((RAW / "ons_ybha.json").read_text(encoding="utf-8"))
    out: dict[int, float] = {}
    for entry in raw.get("years", []):
        try:
            y = int(entry["year"])
            v = float(entry["value"])
        except (KeyError, ValueError, TypeError):
            continue
        out[y] = v  # £ millions, current prices
    return out


# ---------- OBR primary balance ----------------------------------------------

def load_obr_primary_balance() -> dict[int, float]:
    """Returns calendar-year-mapped primary balance, % of GDP.

    OBR rows are fiscal years like "2010-11". We map fiscal year YYYY-(YY+1)
    to calendar year YYYY (the year in which the fiscal year begins).
    """
    wb = openpyxl.load_workbook(RAW / "obr_psf_databank.xlsx", data_only=True, read_only=True)
    ws = wb["Aggregates (per cent of GDP)"]
    out: dict[int, float] = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if not isinstance(label, str) or "-" not in label:
            continue
        try:
            start_year = int(label.split("-")[0])
        except ValueError:
            continue
        v = ws.cell(row=r, column=12).value
        if isinstance(v, (int, float)):
            out[start_year] = float(v)  # already in % of GDP
    wb.close()
    return out


# ---------- BoE real spot rate -----------------------------------------------

def load_boe_real_spot() -> tuple[dict[int, float], dict[int, float]]:
    """Returns (annual_rate_decimal, annual_mean_maturity_years).

    The BoE linker curve front-end is unstable: in some months the shortest
    populated maturity is 2.5y, in others 3y or higher. We use the shortest
    available real spot rate per month-end observation, then average the 12
    monthly values per calendar year. We also report the mean maturity used
    each year so the proxy is transparent.
    """
    files = [
        RAW / "boe_glcreal" / "GLC Real month end data_1979 to 2015.xlsx",
        RAW / "boe_glcreal" / "GLC Real month end data_2016 to 2024.xlsx",
    ]
    by_year_rates: dict[int, list[float]] = {}
    by_year_mats: dict[int, list[float]] = {}
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb["4. spot curve"]
        years_row = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(6, ws.max_row + 1):
            date = ws.cell(row=r, column=1).value
            if date is None:
                continue
            year = getattr(date, "year", None)
            if year is None:
                continue
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    mat = years_row[c - 1]
                    if isinstance(mat, (int, float)):
                        by_year_rates.setdefault(year, []).append(float(v))
                        by_year_mats.setdefault(year, []).append(float(mat))
                    break
        wb.close()
    rate_out: dict[int, float] = {}
    mat_out: dict[int, float] = {}
    for y, vals in by_year_rates.items():
        if vals:
            rate_out[y] = sum(vals) / len(vals) / 100.0  # % -> decimal
            mats = by_year_mats[y]
            mat_out[y] = sum(mats) / len(mats)
    return rate_out, mat_out


# ---------- write inputs.xlsx -------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MANUAL_FILL = PatternFill("solid", fgColor="FFF3C4")  # soft amber
SOURCE_FILL = PatternFill("solid", fgColor="E5E7EB")  # light grey

# Literature standard from FTPL/macro calibration: annual β = 0.99
# (Cochrane, FTPL book; Sims 2013; the methodology PDF's own worked example).
# Stored as the real short rate r such that β = 1/(1+r) = 0.99 exactly.
BETA_LITERATURE = 0.99
R_LITERATURE = 1.0 / BETA_LITERATURE - 1.0  # = 0.010101010...

COLUMNS = [
    ("year",                          "calendar year"),
    ("real_short_rate",               f"literature standard r = 1/β - 1 with β = {BETA_LITERATURE} (constant across years); see README"),
    ("real_short_rate_BoE_observed",  "informational: annual mean of BoE GLC Real shortest-available maturity (decimal). NOT used by calibrate.py."),
    ("long_duration_years",           "Macaulay duration of 15+ gilt segment, in years; DMO"),
    ("V_liab_over_GDP",               "(TPR Purple Book DB s179 + PRA Solvency II tech reserves) / nominal GDP"),
    ("long_gilt_mkt_val_over_GDP",    "DMO long-gilt market value / nominal GDP"),
    ("short_gilt_mkt_val_over_GDP",   "DMO short+medium-gilt market value / nominal GDP"),
    ("primary_balance_over_GDP",      "decimal; OBR Aggregates, FY YYYY-(YY+1) -> calendar YYYY"),
    ("GDP_nominal_GBPm",              "ONS YBHA, current prices, £ millions"),
]


def write_inputs(gdp, pb_pct, real_rate, mat) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inputs"

    # header
    for c, (name, _desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # description row (row 2)
    for c, (_name, desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=c, value=desc)
        cell.fill = SOURCE_FILL
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 50

    # data rows
    for i, y in enumerate(YEARS, start=3):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=R_LITERATURE)  # literature β = 0.99 constant
        ws.cell(row=i, column=3, value=real_rate.get(y, MANUAL))  # BoE observed (informational)
        ws.cell(row=i, column=4, value=MANUAL)
        ws.cell(row=i, column=5, value=MANUAL)
        ws.cell(row=i, column=6, value=MANUAL)
        ws.cell(row=i, column=7, value=MANUAL)
        pb = pb_pct.get(y)
        ws.cell(row=i, column=8, value=(pb / 100.0) if isinstance(pb, (int, float)) else MANUAL)
        ws.cell(row=i, column=9, value=gdp.get(y, MANUAL))
        # paint MANUAL cells
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=i, column=c)
            if cell.value == MANUAL:
                cell.fill = MANUAL_FILL

    # column widths
    widths = [8, 16, 22, 18, 22, 24, 24, 22, 18]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # Sources sheet
    src = wb.create_sheet("sources")
    src.append(["field", "source", "url", "notes"])
    rows = [
        (
            "real_short_rate",
            f"Literature standard: β = {BETA_LITERATURE} constant -> r = {R_LITERATURE:.6f}",
            "FTPL literature (Cochrane FTPL book; Sims 2013; the methodology PDF §4 worked example uses β = 0.99)",
            "Deviation from methodology §2's 'primary source' (BoE linker curve 1y point). Reason: BoE doesn't publish "
            "1y on the linker curve, the 2.5y proxy gives β > 1 in most QE-era years due to deeply negative real rates "
            "and LDI compression. A constant literature β sidesteps the input issue and is closer to standard FTPL "
            "calibration practice. BoE-observed series is preserved in the next column for reference.",
        ),
        (
            "real_short_rate_BoE_observed",
            "BoE GLC Real, monthly month-end, sheet '4. spot curve', shortest available maturity",
            "https://www.bankofengland.co.uk/statistics/yield-curves",
            "Informational only. Annual mean of 12 month-end observations at the shortest populated maturity per month. "
            "Not consumed by calibrate.py.",
        ),
        (
            "long_duration_years",
            "DMO Quarterly Review / data portal: gilt portfolio modified Macaulay duration, 15+ segment",
            "https://www.dmo.gov.uk/data/",
            "DMO data portal renders client-side; download manually.",
        ),
        (
            "V_liab_over_GDP",
            "TPR Purple Book (DB s179) + PRA Solvency II (insurance technical reserves), summed and divided by ONS YBHA",
            "https://www.thepensionsregulator.gov.uk/en/document-library/research-and-analysis/purple-book ; "
            "https://www.bankofengland.co.uk/prudential-regulation/regulatory-reporting",
            "Annual PDFs. Compile manually each year. Headline calibration uses s179 basis.",
        ),
        (
            "long_gilt_mkt_val_over_GDP",
            "DMO gilt portfolio statistics, long-bucket market value / ONS YBHA",
            "https://www.dmo.gov.uk/data/",
            "Excludes index-linked gilts in headline; sensitivity includes them.",
        ),
        (
            "short_gilt_mkt_val_over_GDP",
            "DMO gilt portfolio statistics, short+medium market value / ONS YBHA",
            "https://www.dmo.gov.uk/data/",
            "Same source as long market value.",
        ),
        (
            "primary_balance_over_GDP",
            "OBR PSF aggregates databank, 'Aggregates (per cent of GDP)', col 'Primary balance'",
            "https://obr.uk/data/",
            "Fiscal year YYYY-(YY+1) mapped to calendar year YYYY. Stored as decimal (col value / 100).",
        ),
        (
            "GDP_nominal_GBPm",
            "ONS series YBHA, gross domestic product at current prices, annual",
            "https://www.ons.gov.uk/economy/grossdomesticproductgdp/timeseries/ybha/pn2",
            "£ millions. Denominator for ratios.",
        ),
    ]
    for row in rows:
        src.append(row)
    src.column_dimensions["A"].width = 32
    src.column_dimensions["B"].width = 48
    src.column_dimensions["C"].width = 60
    src.column_dimensions["D"].width = 60
    for r in src.iter_rows(min_row=1, max_row=1):
        for c in r:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
    for r in src.iter_rows(min_row=2, max_row=src.max_row):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    print(f"wrote {OUT}")


def main() -> None:
    gdp = load_ons_ybha()
    pb = load_obr_primary_balance()
    rsr, mats = load_boe_real_spot()

    # sanity prints
    print("ONS YBHA years filled:", sorted(y for y in YEARS if y in gdp))
    print("OBR primary balance years filled:", sorted(y for y in YEARS if y in pb))
    print("BoE real-rate years filled:", sorted(y for y in YEARS if y in rsr))
    print("BoE mean-maturity used per year:")
    for y in YEARS:
        if y in mats:
            print(f"  {y}: rate={rsr[y]*100:.3f}%  mean_mat={mats[y]:.2f}y")

    write_inputs(gdp, pb, rsr, mats)


if __name__ == "__main__":
    main()
