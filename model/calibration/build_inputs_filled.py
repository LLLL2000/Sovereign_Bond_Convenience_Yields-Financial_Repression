"""Build inputs_filled.xlsx — same structure as inputs.xlsx but with the four
MANUAL fields populated by best-guess values plus provenance notes.

Provenance:
  - V_liab_over_GDP numerator
      = Purple Book s179 DB liabilities (PPF 2025 historical table, page 14)
      + BoE Insurance Aggregate technical provisions (Q4, CSV).
      Purple Book years 2010, 2012-2015 LINEARLY INTERPOLATED between the
      reported 2006, 2011, 2016 anchors.
      Insurance TPs pre-2017 EXTRAPOLATED by holding the 2017 ratio of
      insurance-TP / GDP constant backward.
      Denominator = ONS YBHA (already in inputs.xlsx).

  - long_gilt_mkt_val_over_GDP
      Best-guess using a calibration anchored on the DMO Annual Review 2024-25:
      end-March 2025 conventional gilt market value = £1,677bn (gross),
      mapped to a 32% long-share (15+ segment) historical norm. Historical
      values for 2010-2023 are reasoned from the known evolution of UK gilt
      issuance (QE era 2010-2021 saw growing long issuance; 2022-2024
      normalisation). Treat as a STARTING POINT requiring verification.

  - short_gilt_mkt_val_over_GDP
      Same DMO anchor for end-March 2025: 1,677 × 0.68 share for short+medium
      buckets. Earlier years scaled by the same shape.

  - long_duration_years
      Best-guess for 15+ segment modified Macaulay duration. The model's
      deterministic-steady-state D = 1/(1 - βδ); typical UK 15+ segment
      duration over the sample window is 17-22 years (longer in QE/ZLB era,
      shorter post-2022 rate hikes). DMO Annual Review 2024-25 reports
      conventional-gilt portfolio-wide modified duration of 7.84 years at
      end-Mar 2025; the 15+ segment is materially longer. Values below
      are a plausible time path consistent with these endpoints.
"""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
PDFS = HERE / "pdfs"
SRC = HERE / "inputs.xlsx"
OUT_DEFAULT = HERE / "inputs_filled.xlsx"

YEARS = list(range(2010, 2025))


# ---- Purple Book s179 (auto from PDF parse output above) -------------
# Values at 31 March of the stated year; £bn.
# Source: PPF Purple Book 2025 p.14 (Figure 4.2) for actual data points;
# linear interpolation for the missing years.

PB_RAW = {  # actual PB historical-table values
    2006: 792.2,
    2011: 969.7,
    2016: 1563.1,
    2017: 1702.9,
    2018: 1643.8,
    2019: 1628.0,
    2020: 1791.3,
    2021: 1673.8,
    2022: 1473.9,
    2023: 1031.5,  # restated
    2024: 947.9,
}


def fill_purple_book() -> dict[int, tuple[float, str]]:
    """Returns {year: (value_GBPbn, provenance_tag)} for 2010..2024."""
    out: dict[int, tuple[float, str]] = {}
    # known points
    for y, v in PB_RAW.items():
        if 2010 <= y <= 2024:
            out[y] = (v, "PB 2025 p.14 actual")
    # interpolate missing in [2010, 2024]: 2010, 2012, 2013, 2014, 2015
    anchors = sorted(PB_RAW.keys())
    for y in YEARS:
        if y in out:
            continue
        # find bracketing anchors
        lo = max(a for a in anchors if a < y)
        hi = min(a for a in anchors if a > y)
        f = (y - lo) / (hi - lo)
        v = PB_RAW[lo] + f * (PB_RAW[hi] - PB_RAW[lo])
        out[y] = (v, f"linear interp {lo}->{hi}")
    return out


# ---- BoE Insurance Aggregate (parsed from CSV) ----------------------

def fill_insurance_tp_q4() -> dict[int, tuple[float, str]]:
    """Q4 Technical Provisions in £bn; 2017Q4..2024Q4 from CSV.
    Pre-2017: extrapolate by holding 2017 ratio TP/GDP constant.
    """
    tps: dict[int, float] = {}
    csv_path = PDFS / "boe_insurance_aggregate.csv"
    by_qtr: dict[str, dict[str, float]] = defaultdict(dict)
    with csv_path.open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["Chart Feature 1"].strip() != "Technical Provisions":
                continue
            period = row["Reporting Period"]
            perm = row["Permissions"]
            try:
                v = float(row["GBP Value"]) if row["GBP Value"] else None
            except ValueError:
                v = None
            if v is not None:
                by_qtr[period][perm] = v
    # take Q4 each year, sum Life + Non-Life
    for q, d in by_qtr.items():
        if not q.endswith("Q4"):
            continue
        year = int(q[:4])
        tps[year] = (d.get("Life", 0.0) + d.get("Non-Life", 0.0)) / 1e9  # to £bn

    out: dict[int, tuple[float, str]] = {}
    for y in YEARS:
        if y in tps:
            out[y] = (tps[y], "BoE Insurance Aggregate Q4 actual")
    return out, tps


# ---- ONS YBHA already saved; reload as £m -----------------------------

def load_gdp() -> dict[int, float]:
    import json
    raw = (RAW / "ons_ybha.json").read_text(encoding="utf-8")
    j = __import__("json").loads(raw)
    out: dict[int, float] = {}
    for e in j.get("years", []):
        try:
            out[int(e["year"])] = float(e["value"])
        except (KeyError, ValueError, TypeError):
            pass
    return out


# ---- DMO best-guess values (anchored on Annual Review 2024-25 p.54) ---
# Conventional gilt MV (end-Mar 2025) = £1,677bn gross  -> 58% of GDP
# Assumed long-share (15+ of conventional) = 32% historical norm
# Time path reflects: 2010-2014 build-up, 2015-2020 QE expansion,
# 2021-2022 stabilisation/hikes, 2023-2024 normalisation.
# All numbers in £bn (will be divided by GDP to produce the ratio).

DMO_LONG_MV = {  # £bn, conventional 15+ market value, end-of-year approx
    2010: 320, 2011: 360, 2012: 410, 2013: 430, 2014: 470,
    2015: 490, 2016: 540, 2017: 555, 2018: 565, 2019: 580,
    2020: 620, 2021: 640, 2022: 525, 2023: 525, 2024: 540,
}

DMO_SHORT_MED_MV = {  # £bn, conventional 0-15yr market value
    2010: 680, 2011: 740, 2012: 830, 2013: 890, 2014: 940,
    2015: 980, 2016: 1080, 2017: 1100, 2018: 1140, 2019: 1170,
    2020: 1320, 2021: 1370, 2022: 1140, 2023: 1135, 2024: 1140,
}

DMO_DURATION_15PLUS = {  # years, 15+ segment Macaulay duration
    2010: 17.0, 2011: 18.5, 2012: 19.5, 2013: 19.0, 2014: 19.8,
    2015: 20.5, 2016: 21.5, 2017: 20.5, 2018: 20.0, 2019: 21.5,
    2020: 22.5, 2021: 21.0, 2022: 17.5, 2023: 17.5, 2024: 17.8,
}


# ---- writer ---------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GUESS_FILL = PatternFill("solid", fgColor="DBEAFE")  # soft blue = best-guess
DATA_FILL = PatternFill("solid", fgColor="D1FAE5")   # soft green = actual data
INTERP_FILL = PatternFill("solid", fgColor="FEF3C7") # soft amber = interpolated
SOURCE_FILL = PatternFill("solid", fgColor="E5E7EB")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default=str(OUT_DEFAULT))
    args = ap.parse_args()
    OUT = Path(args.output)

    gdp = load_gdp()
    pb = fill_purple_book()
    ins_q4, ins_raw = fill_insurance_tp_q4()
    # pre-2017 insurance extrapolation: hold 2017 ratio constant
    if 2017 in ins_raw and 2017 in gdp:
        ratio_2017 = ins_raw[2017] * 1e3 / gdp[2017]  # £bn / £m -> dimensionless
        for y in YEARS:
            if y not in ins_q4 and y in gdp:
                ins_q4[y] = (ratio_2017 * gdp[y] / 1e3, "extrapolated @ 2017 TP/GDP ratio")

    # Load the inputs.xlsx so we keep the auto-filled columns
    src_wb = openpyxl.load_workbook(SRC, data_only=True)
    src_ws = src_wb["inputs"]
    headers = [src_ws.cell(row=1, column=c).value for c in range(1, src_ws.max_column + 1)]
    src_data: list[dict] = []
    for r in range(3, src_ws.max_row + 1):
        row = {h: src_ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
        src_data.append(row)
    src_wb.close()

    # build the output
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inputs"

    # header rows (same as inputs.xlsx + provenance column at end)
    out_headers = headers + ["provenance"]
    for c, h in enumerate(out_headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # description row 2 — copy from src then add provenance
    src_desc = [src_ws.cell(row=2, column=c).value for c in range(1, src_ws.max_column + 1)] \
        if False else None
    # Just leave the description row sparse — readers can reference inputs.xlsx
    desc = {
        "year": "calendar year",
        "real_short_rate": "literature β = 0.99 constant (r = 1/β - 1)",
        "real_short_rate_BoE_observed": "informational only; BoE GLC Real annual mean",
        "real_rate_mean_maturity_y": "informational only; BoE GLC Real annual mean",
        "long_duration_years": "15+ Macaulay duration; best-guess (see provenance sheet)",
        "V_liab_over_GDP": "(PB s179 + BoE insurance TP) / GDP",
        "long_gilt_mkt_val_over_GDP": "best-guess: long share of conv. gilt MV",
        "short_gilt_mkt_val_over_GDP": "best-guess: short+medium share of conv. gilt MV",
        "primary_balance_over_GDP": "decimal; OBR Aggregates",
        "GDP_nominal_GBPm": "ONS YBHA",
        "provenance": "per-cell tag",
    }
    for c, h in enumerate(out_headers, start=1):
        cell = ws.cell(row=2, column=c, value=desc.get(h, ""))
        cell.fill = SOURCE_FILL
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 40

    # data rows
    legend = []  # collect for provenance sheet
    for i, row in enumerate(src_data, start=3):
        y = int(row["year"])
        ws.cell(row=i, column=1, value=y)
        # Inherit literature β real-rate from inputs.xlsx (col 2)
        # and BoE observed series (col 3, informational).
        ws.cell(row=i, column=2, value=row["real_short_rate"])
        ws.cell(row=i, column=3, value=row.get("real_short_rate_BoE_observed", row.get("real_rate_mean_maturity_y")))

        # 4: long_duration_years
        ws.cell(row=i, column=4, value=DMO_DURATION_15PLUS[y]).fill = GUESS_FILL

        # 5: V_liab/GDP = (Purple Book + Insurance TP) / GDP
        pb_v, pb_tag = pb[y]
        ins_v, ins_tag = ins_q4[y]
        gdp_y = gdp[y]
        v_liab = (pb_v + ins_v) * 1e3 / gdp_y  # both in £bn; gdp in £m
        c5 = ws.cell(row=i, column=5, value=v_liab)
        c5.fill = DATA_FILL if (pb_tag == "PB 2025 p.14 actual" and ins_tag == "BoE Insurance Aggregate Q4 actual") else INTERP_FILL

        # 6: long gilt MV / GDP
        c6 = ws.cell(row=i, column=6, value=DMO_LONG_MV[y] * 1e3 / gdp_y)
        c6.fill = GUESS_FILL

        # 7: short+medium gilt MV / GDP
        c7 = ws.cell(row=i, column=7, value=DMO_SHORT_MED_MV[y] * 1e3 / gdp_y)
        c7.fill = GUESS_FILL

        # 8-9: keep original
        ws.cell(row=i, column=8, value=row["primary_balance_over_GDP"])
        ws.cell(row=i, column=9, value=row["GDP_nominal_GBPm"])

        # 10: provenance string
        tag = f"PB:{pb_tag} | INS:{ins_tag} | DMO:best-guess"
        ws.cell(row=i, column=10, value=tag).alignment = Alignment(wrap_text=True, vertical="top")

        legend.append((y, pb_v, pb_tag, ins_v, ins_tag, DMO_LONG_MV[y], DMO_SHORT_MED_MV[y], DMO_DURATION_15PLUS[y]))

    # widths
    widths = [8, 14, 18, 18, 18, 24, 24, 22, 18, 70]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # provenance sheet
    pr = wb.create_sheet("provenance")
    pr.append(["year", "PB_s179_GBPbn", "PB_provenance", "Insurance_TP_Q4_GBPbn", "Insurance_provenance",
               "DMO_long_MV_GBPbn", "DMO_shortmed_MV_GBPbn", "DMO_15plus_duration_y"])
    for row in legend:
        pr.append(list(row))
    for cell in pr[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    pr.column_dimensions["A"].width = 6
    for col in "BCDEFGH":
        pr.column_dimensions[col].width = 22

    # legend sheet
    lg = wb.create_sheet("legend")
    lg.append(["color", "meaning"])
    lg.append(["green", "actual reported value from primary source"])
    lg.cell(row=2, column=1).fill = DATA_FILL
    lg.append(["amber", "interpolated or extrapolated between actual data points"])
    lg.cell(row=3, column=1).fill = INTERP_FILL
    lg.append(["blue", "best-guess by Claude; cross-check before thesis use"])
    lg.cell(row=4, column=1).fill = GUESS_FILL
    for cell in lg[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    lg.column_dimensions["A"].width = 12
    lg.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
