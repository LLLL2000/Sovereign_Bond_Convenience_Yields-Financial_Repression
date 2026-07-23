"""Solve the augmented FTPL identity year by year (§3.4, Listing 1).

Reads ./inputs.xlsx, applies the closed-form algebra at each calendar year
2010..2024, asserts sanity conditions, and writes ./outputs.xlsx.

Rows where any required input is "MANUAL" (or missing) are skipped and listed
in the missing-inputs sheet.
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent

MANUAL = "MANUAL"

REQUIRED = [
    "real_short_rate",
    "long_duration_years",
    "V_liab_over_GDP",
    "long_gilt_mkt_val_over_GDP",
    "short_gilt_mkt_val_over_GDP",
    "primary_balance_over_GDP",
]

OUTPUT_COLS = [
    "year",
    "beta",
    "delta",
    "Q_L_fund",
    "v",
    "R",
    "rent_share",
    "omega",
    "Q_L",
    "lambda_L",
    "b_L",
    "yield_wedge",
    "zeta",
]


def read_inputs(path: Path) -> list[dict]:
    """Return one dict per year. Missing inputs left as None."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["inputs"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        row = {h: ws.cell(row=r, column=c + 1).value for c, h in enumerate(headers)}
        # normalize MANUAL/None
        for k in REQUIRED:
            v = row.get(k)
            if v == MANUAL or v is None:
                row[k] = None
            elif isinstance(v, str):
                try:
                    row[k] = float(v)
                except ValueError:
                    row[k] = None
            else:
                row[k] = float(v)
        rows.append(row)
    wb.close()
    return rows


def calibrate_year(row: dict) -> dict | None:
    r = row["real_short_rate"]
    D = row["long_duration_years"]
    V_liab = row["V_liab_over_GDP"]
    bL_mkt = row["long_gilt_mkt_val_over_GDP"]
    bS_mkt = row["short_gilt_mkt_val_over_GDP"]
    s = row["primary_balance_over_GDP"]
    if any(x is None for x in [r, D, V_liab, bL_mkt, bS_mkt, s]):
        return None

    beta = 1.0 / (1.0 + r)
    delta = (1.0 - 1.0 / D) / beta
    Q_L_fund = beta / (1.0 - beta * delta)

    v = bS_mkt + bL_mkt
    R = (1.0 - beta) * v - s
    rent_share = R / bL_mkt
    # omega may go negative if (1 - rent_share) <= 0 or rent_share < 0;
    # this is informative — we report it as-is.
    if (1.0 - rent_share) == 0:
        omega = float("inf")
    else:
        omega = rent_share / (1.0 - rent_share)

    Q_L = (1.0 + omega) * Q_L_fund
    lambda_L = bL_mkt / V_liab
    b_L_model = bL_mkt / Q_L if Q_L != 0 else float("nan")
    yield_wedge = (1.0 / Q_L_fund) - (1.0 / Q_L) if Q_L != 0 else float("nan")
    zeta = (R / (1.0 - beta)) / v if (1.0 - beta) != 0 and v != 0 else float("nan")

    return {
        "year": int(row["year"]),
        "beta": beta,
        "delta": delta,
        "Q_L_fund": Q_L_fund,
        "v": v,
        "R": R,
        "rent_share": rent_share,
        "omega": omega,
        "Q_L": Q_L,
        "lambda_L": lambda_L,
        "b_L": b_L_model,
        "yield_wedge": yield_wedge,
        "zeta": zeta,
    }


def sanity_checks(results: list[dict]) -> list[tuple[int, str, bool, str]]:
    """Return (year, check_name, passed, detail) tuples for the assertions in §5."""
    out: list[tuple[int, str, bool, str]] = []
    for r in results:
        y = r["year"]
        out.append((y, "omega>0 (Assumption 1)", r["omega"] > 0, f"omega={r['omega']:.4f}"))
        out.append((y, "lambda_L in (0,1)", 0.0 < r["lambda_L"] < 1.0, f"lambda_L={r['lambda_L']:.4f}"))
        bd = r["beta"] * r["delta"]
        out.append((y, "beta*delta<1 (price finite)", bd < 1.0, f"beta*delta={bd:.6f}"))
    return out


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FAIL_FILL = PatternFill("solid", fgColor="FECACA")
PASS_FILL = PatternFill("solid", fgColor="D1FAE5")


def write_outputs(out_path: Path, rows: list[dict], results: list[dict], checks, missing: list[tuple[int, list[str]]]) -> None:
    wb = openpyxl.Workbook()

    # echo inputs
    ws_in = wb.active
    ws_in.title = "inputs (echo)"
    if rows:
        headers = list(rows[0].keys())
        ws_in.append(headers)
        for r in rows:
            ws_in.append([r.get(h) for h in headers])
        for cell in ws_in[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    # outputs
    ws_out = wb.create_sheet("calibration")
    ws_out.append(OUTPUT_COLS)
    for r in results:
        ws_out.append([r.get(k) for k in OUTPUT_COLS])
    for cell in ws_out[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    # number format
    for c in range(2, len(OUTPUT_COLS) + 1):
        for cell in ws_out.iter_cols(min_col=c, max_col=c, min_row=2, max_row=ws_out.max_row):
            for k in cell:
                k.number_format = "0.000000"
    ws_out.column_dimensions["A"].width = 8
    for letter in "BCDEFGHIJKLM":
        ws_out.column_dimensions[letter].width = 13

    # sanity
    ws_chk = wb.create_sheet("sanity")
    ws_chk.append(["year", "check", "passed", "detail"])
    for cell in ws_chk[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for y, name, ok, detail in checks:
        ws_chk.append([y, name, "PASS" if ok else "FAIL", detail])
        cell = ws_chk.cell(row=ws_chk.max_row, column=3)
        cell.fill = PASS_FILL if ok else FAIL_FILL
    ws_chk.column_dimensions["A"].width = 8
    ws_chk.column_dimensions["B"].width = 32
    ws_chk.column_dimensions["C"].width = 10
    ws_chk.column_dimensions["D"].width = 28

    # missing
    ws_miss = wb.create_sheet("missing inputs")
    ws_miss.append(["year", "missing fields"])
    for cell in ws_miss[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for y, fields in missing:
        ws_miss.append([y, ", ".join(fields)])
    ws_miss.column_dimensions["A"].width = 8
    ws_miss.column_dimensions["B"].width = 80
    for r in ws_miss.iter_rows(min_row=2, max_row=ws_miss.max_row):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=str(HERE / "inputs.xlsx"))
    ap.add_argument("--output", default=None, help="defaults to outputs.xlsx next to input")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output) if args.output else in_path.with_name(in_path.stem.replace("inputs", "outputs") + ".xlsx")

    rows = read_inputs(in_path)
    results: list[dict] = []
    missing: list[tuple[int, list[str]]] = []

    for row in rows:
        miss = [k for k in REQUIRED if row.get(k) is None]
        if miss:
            missing.append((int(row["year"]), miss))
            continue
        out = calibrate_year(row)
        if out is None:
            missing.append((int(row["year"]), REQUIRED))
            continue
        results.append(out)

    checks = sanity_checks(results)
    write_outputs(out_path, rows, results, checks, missing)

    print(f"input:  {in_path}")
    print(f"output: {out_path}")
    print(f"calibrated years: {[r['year'] for r in results]}")
    print(f"skipped (missing inputs): {[y for y, _ in missing]}")
    if results:
        print()
        print("headline implied wedges:")
        for r in results:
            print(f"  {r['year']}: omega={r['omega']:.4f}  Q_L={r['Q_L']:.3f}  lambda_L={r['lambda_L']:.4f}")
    failed = [c for c in checks if not c[2]]
    if failed:
        print(f"\nSANITY: {len(failed)} check(s) failed:")
        for y, name, _, detail in failed:
            print(f"  year {y}: {name} -- {detail}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
