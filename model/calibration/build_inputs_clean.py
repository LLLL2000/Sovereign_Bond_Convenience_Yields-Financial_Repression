"""Build inputs_clean.xlsx — the restricted, clean-input sample for the
dissertation's §5.2 headline omega.

What's clean:
  - real_short_rate:           literature beta = 0.99 (constant, by convention)
  - primary_balance_over_GDP:  OBR PSF databank, all years (auto-parsed)
  - GDP_nominal_GBPm:          ONS YBHA, all years (auto-parsed)
  - V_liab numerator:
      PB s179:           ACTUAL from PPF 2025 p.14 for 2011, 2016-2024
      Insurance TP Q4:   ACTUAL from BoE Insurance Aggregate CSV for 2017-2024
      Joint-actual:      2017-2024
  - long_duration_years (portfolio-wide proxy):
      portfolio-wide modified duration of conventional gilts at fiscal year-end
      (DMO Annual Review). NOT the 15+ segment specifically — the DMO does not
      publish duration broken out by maturity bucket. We treat this as the
      cleanest available proxy and document the deviation in the report.
      ACTUAL for calendar 2021 (end-Mar 2022) through 2024 (end-Mar 2025).
  - long_gilt_mkt_val_over_GDP + short_gilt_mkt_val_over_GDP:
      total conv. gilt market value (DMO AR) / GDP, split by long-share.
      The long-share is not directly observable in any DMO publication; we
      derive it from the nominal-weighted Short/Medium/Long classification in
      the QR Apr-Jun 2025 gilts-in-issue list (page 7). This is an audit-able
      assumption, used as the headline split, with a sensitivity band reported.

Joint clean window for ALL inputs: 2021-2024 (4 years).

The pipeline writes a row only for these four years; calibrate.py treats the
xlsx exactly like inputs_filled_v2.xlsx.

Run:  python build_inputs_clean.py
Outputs: inputs_clean.xlsx
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
PDFS = HERE / "pdfs"
OUT = HERE / "inputs_clean.xlsx"

CLEAN_YEARS = [2021, 2022, 2023, 2024]

# ---- DMO Annual Review snapshots (fiscal year-end, gross values) ------
# Calendar year YYYY <-> fiscal-year-end 31 March (YYYY+1).
# Sources:
#   GAR 2022-23 (https://www.dmo.gov.uk/media/tfidb5fy/gar2023.pdf), Tables 19-20
#   GAR 2023-24 (https://www.dmo.gov.uk/media/5rqb2scf/gar2024_final.pdf), Tables 19-20
#   GAR 2024-25 (pdfs/dmo_annual_review_2024_25.pdf), Table 20

DMO_CONV_MV_GBPBN = {
    2021: 1748.0,   # 31 Mar 2022; gross, conventional gilts
    2022: 1547.0,   # 31 Mar 2023
    2023: 1612.0,   # 31 Mar 2024
    2024: 1677.0,   # 31 Mar 2025
}

DMO_CONV_MOD_DURATION_Y = {  # market value-weighted, conventional gilts, portfolio-wide
    2021: 11.47,   # 31 Mar 2022
    2022:  9.17,   # 31 Mar 2023
    2023:  8.39,   # 31 Mar 2024
    2024:  7.84,   # 31 Mar 2025
}

# ---- Long-share derivation: QR Apr-Jun 2025, gilts-in-issue list (p.7) ---
# DMO's standard maturity classification (Short / Medium / Long) applied to
# each conventional gilt in issue at 30-Jun-2025. We sum nominal £mn by bucket
# to get a nominal-weighted long share. This is a defensible alternative to
# the README's flat 32% best-guess. MV-weighted shares may differ slightly;
# longer gilts trade at deeper discounts to par when yields rise, so the
# nominal-weighted share is a mild over-estimate of the MV-weighted share.

QR_GILTS_30JUN2025 = {
    "short": [
        ("2% 2025",     40343.0), ("3.5% 2025",   36016.3), ("0.125% 2026", 41177.7),
        ("1.5% 2026",   44673.7), ("0.375% 2026", 33660.6), ("4.125% 2027", 33031.1),
        ("3.75% 2027",  37352.7), ("1.25% 2027",  41947.8), ("4.25% 2027",  33776.8),
        ("0.125% 2028", 35260.3), ("4.375% 2028", 30742.4), ("4.5% 2028",   35217.0),
        ("1.625% 2028", 38743.3), ("6% 2028",     20730.6), ("0.5% 2029",   29597.5),
        ("4.125% 2029", 37987.0), ("0.875% 2029", 44643.1), ("4.375% 2030", 29824.2),
        ("0.375% 2030", 39816.7), ("4.75% 2030",  43823.4), ("0.25% 2031",  41587.9),
        ("4% 2031",     32623.4), ("1% 2032",     36801.4), ("4.25% 2032",  41276.1),
    ],
    "medium": [
        ("3.25% 2033",  34175.1), ("0.875% G 2033", 39783.0), ("4.625% 2034", 32512.3),
        ("4.25% 2034",  36102.6), ("4.5% 2034",   37112.3), ("4.5% 2035",   30365.4),
        ("0.625% 2035", 35633.9), ("4.25% 2036",  32424.9), ("1.75% 2037",  32718.3),
        ("3.75% 2038",  32888.6), ("4.75% 2038",  27455.9), ("1.125% 2039", 24950.6),
        ("4.25% 2039",  24802.8), ("4.375% 2040", 27363.5),
    ],
    "long": [
        ("4.25% 2040",  27069.1), ("1.25% 2041",  34610.2), ("4.5% 2042",   29019.3),
        ("4.75% 2043",  33065.0), ("3.25% 2044",  29790.6), ("3.5% 2045",   30093.4),
        ("0.875% 2046", 23521.6), ("4.25% 2046",  26714.2), ("1.5% 2047",   26325.4),
        ("1.75% 2049",  30932.8), ("4.25% 2049",  21541.4), ("0.625% 2050", 32663.9),
        ("1.25% 2051",  29538.7), ("3.75% 2052",  25703.1), ("1.5% G 2053", 26928.0),
        ("3.75% 2053",  28288.0), ("4.375% 2054", 32127.9), ("1.625% 2054", 25158.1),
        ("4.25% 2055",  28320.2), ("5.375% 2056",  4000.0), ("1.75% 2057",  31462.5),
        ("4% 2060",     25722.4), ("0.5% 2061",   26493.2), ("4% 2063",     18538.5),
        ("2.5% 2065",   21057.9), ("3.5% 2068",   21230.4), ("1.625% 2071", 24725.3),
        ("1.125% 2073", 11111.0),
    ],
}


def derive_long_share() -> float:
    """Long-share of conventional gilt nominal, 30-Jun-2025."""
    totals = {k: sum(v for _, v in lst) for k, lst in QR_GILTS_30JUN2025.items()}
    grand = sum(totals.values())
    return totals["long"] / grand


# ---- ONS YBHA (auto) ------------------------------------------------

def load_gdp() -> dict[int, float]:
    raw = json.loads((RAW / "ons_ybha.json").read_text(encoding="utf-8"))
    out: dict[int, float] = {}
    for e in raw.get("years", []):
        try:
            out[int(e["year"])] = float(e["value"])
        except (KeyError, ValueError, TypeError):
            pass
    return out


# ---- OBR PSF databank (auto) ---------------------------------------

def load_obr_primary_balance() -> dict[int, float]:
    wb = openpyxl.load_workbook(RAW / "obr_psf_databank.xlsx", data_only=True, read_only=True)
    ws = wb["Aggregates (per cent of GDP)"]
    out: dict[int, float] = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if not isinstance(label, str) or "-" not in label:
            continue
        try:
            start_year = int(label.split("-")[0])
        except ValueError:
            continue
        v = ws.cell(row=r, column=12).value
        if isinstance(v, (int, float)):
            out[start_year] = float(v) / 100.0  # decimal
    wb.close()
    return out


# ---- Purple Book s179 (auto from PPF 2025 p.14) ---------------------
# ACTUAL values only — no interpolation.
PB_ACTUAL_GBPBN = {
    2011: 969.7,
    2016: 1563.1,
    2017: 1702.9,
    2018: 1643.8,
    2019: 1628.0,
    2020: 1791.3,
    2021: 1673.8,
    2022: 1473.9,
    2023: 1031.5,
    2024:  947.9,
}


# ---- BoE Insurance Aggregate Q4 (auto from CSV) ---------------------

def load_insurance_tp_q4() -> dict[int, float]:
    """Q4 Technical Provisions in £bn; only actual reported years (2017+)."""
    by_qtr: dict[str, dict[str, float]] = defaultdict(dict)
    with (PDFS / "boe_insurance_aggregate.csv").open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["Chart Feature 1"].strip() != "Technical Provisions":
                continue
            period = row["Reporting Period"]
            perm = row["Permissions"]
            try:
                v = float(row["GBP Value"]) if row["GBP Value"] else None
            except ValueError:
                v = None
            if v is not None:
                by_qtr[period][perm] = v
    out: dict[int, float] = {}
    for q, d in by_qtr.items():
        if not q.endswith("Q4"):
            continue
        out[int(q[:4])] = (d.get("Life", 0.0) + d.get("Non-Life", 0.0)) / 1e9
    return out


# ---- writer ---------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DATA_FILL = PatternFill("solid", fgColor="D1FAE5")     # green: clean / actual
ASSUM_FILL = PatternFill("solid", fgColor="FFF7CC")    # pale yellow: assumption (long-share)
SOURCE_FILL = PatternFill("solid", fgColor="E5E7EB")

BETA = 0.99
R_LIT = 1.0 / BETA - 1.0


def main() -> None:
    long_share = derive_long_share()
    print(f"Derived long-share (QR Apr-Jun 2025, nominal-weighted): {long_share:.4f}")

    gdp = load_gdp()
    obr = load_obr_primary_balance()
    insurance = load_insurance_tp_q4()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "inputs"

    headers = [
        "year",
        "real_short_rate",
        "real_short_rate_BoE_observed",   # left blank (informational only; unused)
        "long_duration_years",
        "V_liab_over_GDP",
        "long_gilt_mkt_val_over_GDP",
        "short_gilt_mkt_val_over_GDP",
        "primary_balance_over_GDP",
        "GDP_nominal_GBPm",
        "provenance",
    ]
    descs = {
        "real_short_rate": f"literature beta = {BETA} (r = 1/beta - 1)",
        "real_short_rate_BoE_observed": "blank in clean run; BoE proxy unused",
        "long_duration_years": "portfolio-wide mod. duration, conv. gilts, DMO AR (clean PROXY for 15+)",
        "V_liab_over_GDP": "(PB s179 actual + BoE Insurance TP Q4 actual) / GDP",
        "long_gilt_mkt_val_over_GDP": f"{long_share:.3f} x (total conv gilt MV / GDP); split = QR Jun-2025 nominal",
        "short_gilt_mkt_val_over_GDP": f"{1-long_share:.3f} x (total conv gilt MV / GDP)",
        "primary_balance_over_GDP": "decimal; OBR Aggregates",
        "GDP_nominal_GBPm": "ONS YBHA",
        "provenance": "per-cell tag",
        "year": "calendar year",
    }
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=descs.get(h, ""))
        cell.fill = SOURCE_FILL
        cell.font = Font(italic=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 50

    legend: list[tuple] = []

    for i, y in enumerate(CLEAN_YEARS, start=3):
        conv_mv_bn = DMO_CONV_MV_GBPBN[y]
        gdp_m = gdp[y]
        v_total = conv_mv_bn * 1e3 / gdp_m
        bL = long_share * v_total
        bS = (1.0 - long_share) * v_total

        v_liab_num_bn = PB_ACTUAL_GBPBN[y] + insurance[y]
        v_liab_over_gdp = v_liab_num_bn * 1e3 / gdp_m

        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=R_LIT)
        ws.cell(row=i, column=3, value=None)
        c4 = ws.cell(row=i, column=4, value=DMO_CONV_MOD_DURATION_Y[y]); c4.fill = DATA_FILL
        c5 = ws.cell(row=i, column=5, value=v_liab_over_gdp); c5.fill = DATA_FILL
        c6 = ws.cell(row=i, column=6, value=bL); c6.fill = ASSUM_FILL
        c7 = ws.cell(row=i, column=7, value=bS); c7.fill = ASSUM_FILL
        ws.cell(row=i, column=8, value=obr[y])
        ws.cell(row=i, column=9, value=gdp_m)
        tag = (
            f"DMO conv MV: GAR @ 31-Mar-{y+1} £{conv_mv_bn:.0f}bn | "
            f"Duration: GAR portfolio-wide {DMO_CONV_MOD_DURATION_Y[y]:.2f}y | "
            f"PB s179: PB 2025 p.14 actual £{PB_ACTUAL_GBPBN[y]:.1f}bn | "
            f"InsTP: BoE Aggregate Q4 actual £{insurance[y]:.1f}bn | "
            f"Long-share: QR Jun-2025 nominal {long_share:.3f}"
        )
        ws.cell(row=i, column=10, value=tag).alignment = Alignment(wrap_text=True, vertical="top")

        legend.append((y, conv_mv_bn, DMO_CONV_MOD_DURATION_Y[y],
                       PB_ACTUAL_GBPBN[y], insurance[y], long_share,
                       bL, bS, v_liab_over_gdp, obr[y], gdp_m))

    widths = [8, 14, 18, 18, 18, 22, 22, 20, 18, 80]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # provenance sheet
    pr = wb.create_sheet("provenance")
    pr.append([
        "year", "DMO_conv_MV_GBPbn", "DMO_conv_modDuration_y",
        "PB_s179_GBPbn_actual", "Insurance_TP_Q4_GBPbn_actual",
        "long_share_QR_Jun2025", "bL_over_GDP", "bS_over_GDP",
        "V_liab_over_GDP", "primary_balance_over_GDP", "GDP_GBPm",
    ])
    for row in legend:
        pr.append(list(row))
    for cell in pr[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    pr.column_dimensions["A"].width = 6
    for col in "BCDEFGHIJK":
        pr.column_dimensions[col].width = 22

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
